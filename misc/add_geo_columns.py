# ──────────────────────────────────────────────────────────────────────────────
# add_geo_columns.py  --  PCT GeoGuesser
#
# Fills in the geographic columns of photos.csv: lat, lon, and map.
#
# Each photo row only knows its trail mile (the NoBo "mile" column). This script
# turns that mile into an approximate latitude/longitude by interpolating against
# a reference table of ~2,900 GPS waypoints that span the whole trail, then builds
# a ready-to-open PCTA interactive-map URL centered on that point.
#
#   lat  : interpolated latitude  (decimal degrees, ~6 dp)
#   lon  : interpolated longitude (decimal degrees, ~6 dp)
#   map  : PCTA sidebar-map deep link, centered on (lon, lat) at zoom level 14
#
# Source of GPS data:
#   misc/halfmiles_pct_notes_gps and miles.xlsx
#   Relevant columns: pct_mile_23 (NoBo mile), lat, lon.
#   This third-party spreadsheet is gitignored (not ours to redistribute), so the
#   script only runs on a machine that has the file sitting next to it in misc/.
#   Requires openpyxl:  pip install openpyxl
#
# Behaviour:
#   By default only fills rows whose lat/lon/map are EMPTY, so any value you have
#   hand-tuned in photos.csv is left untouched (same "manual edits win" rule that
#   generate_photo_csv.py follows). Pass --force to recompute every row from the
#   spreadsheet instead.
#
# Typical flow when adding new photos:
#   1. python3 misc/generate_photo_csv.py   (adds new rows; lat/lon/map left blank)
#   2. python3 misc/add_geo_columns.py      (fills the blanks)
#
# Run:
#   python3 misc/add_geo_columns.py            # fill blanks only
#   python3 misc/add_geo_columns.py --force    # recompute all rows
# ──────────────────────────────────────────────────────────────────────────────

import os
import csv
import sys
import bisect

import openpyxl

_HERE     = os.path.dirname(os.path.abspath(__file__))
CSV_PATH  = os.path.join(_HERE, "photos.csv")
XLSX_PATH = os.path.join(_HERE, "halfmiles_pct_notes_gps and miles.xlsx")

# PCTA interactive sidebar map. The &center=<lon>,<lat>&level=<zoom> params drop
# you straight onto the photo's location. Level 14 is roughly trail-segment zoom.
MAP_BASE  = "https://pcta.maps.arcgis.com/apps/instant/sidebar/index.html?appid=3b1817932adf42009f30b6b38828212e"
MAP_LEVEL = 14

FORCE = "--force" in sys.argv[1:]


def load_reference():
    """Return a list of (mile, lat, lon) waypoints, sorted by mile.

    Only rows that have all three values are kept. The spreadsheet has many
    note-only rows (water reports, directions) with no coordinates; those are
    skipped so they can't poison the interpolation.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    header  = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    mi_idx  = header.index("pct_mile_23")
    lat_idx = header.index("lat")
    lon_idx = header.index("lon")

    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mile, lat, lon = row[mi_idx], row[lat_idx], row[lon_idx]
        if mile is not None and lat is not None and lon is not None:
            points.append((float(mile), float(lat), float(lon)))
    wb.close()

    points.sort()
    return points


def interpolate(points, miles, target):
    """Linearly interpolate (lat, lon) for a trail mile.

    `miles` is the pre-extracted sorted list of reference miles (so we don't
    rebuild it on every call). Targets past either end clamp to the nearest
    endpoint rather than extrapolating into nonsense.
    """
    idx = bisect.bisect_left(miles, target)
    if idx == 0:
        return round(points[0][1], 6), round(points[0][2], 6)
    if idx >= len(points):
        return round(points[-1][1], 6), round(points[-1][2], 6)

    lo, hi = points[idx - 1], points[idx]
    span   = hi[0] - lo[0]
    t      = 0.0 if span == 0 else (target - lo[0]) / span
    lat = lo[1] + t * (hi[1] - lo[1])
    lon = lo[2] + t * (hi[2] - lo[2])
    return round(lat, 6), round(lon, 6)


def map_url(lat, lon):
    # center expects lon,lat order (x,y), not lat,lon.
    return f"{MAP_BASE}&center={lon},{lat}&level={MAP_LEVEL}"


def main():
    if not os.path.exists(XLSX_PATH):
        sys.exit(f"GPS spreadsheet not found: {XLSX_PATH}\n"
                 f"This file lives under the gitignored img/ dir; copy it in first.")

    points = load_reference()
    miles  = [p[0] for p in points]
    print(f"Loaded {len(points)} GPS reference points "
          f"(miles {points[0][0]:.1f} to {points[-1][0]:.1f}).")

    with open(CSV_PATH, newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    # Make sure the three columns exist in the header, appended after the rest.
    for col in ("lat", "lon", "map"):
        if col not in fieldnames:
            fieldnames.append(col)

    filled = 0
    for r in rows:
        has_geo = r.get("lat") and r.get("lon") and r.get("map")
        if has_geo and not FORCE:
            continue  # respect existing / hand-tuned values

        try:
            mile = float(r["mile"])
        except (KeyError, ValueError):
            continue  # a row with no usable mile can't be placed

        lat, lon = interpolate(points, miles, mile)
        r["lat"] = str(lat)
        r["lon"] = str(lon)
        r["map"] = map_url(lat, lon)
        filled += 1

    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    verb = "recomputed" if FORCE else "filled"
    print(f"photos.csv: {verb} lat/lon/map for {filled} of {len(rows)} rows.")


if __name__ == "__main__":
    main()
